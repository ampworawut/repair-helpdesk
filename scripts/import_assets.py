#!/usr/bin/env python3
"""
นำเข้าข้อมูลเครื่องคอมพิวเตอร์จาก Excel → Supabase
ใช้ Com name เป็น asset_code ตาม requirement

วิธีใช้:
  python3 scripts/import_assets.py <path/to/excel.xlsx>

.env ต้องมี:
  NEXT_PUBLIC_SUPABASE_URL=...
  SUPABASE_SERVICE_ROLE_KEY=...  (ใช้ service_role key สำหรับ bypass RLS)
"""

import sys
import os
import openpyxl
from datetime import datetime
from supabase import create_client, Client

SUPABASE_URL = os.environ.get("NEXT_PUBLIC_SUPABASE_URL", "")
SUPABASE_KEY = os.environ.get("SUPABASE_SERVICE_ROLE_KEY", "")

if not SUPABASE_URL or not SUPABASE_KEY:
    print("❌ ต้องตั้งค่า NEXT_PUBLIC_SUPABASE_URL และ SUPABASE_SERVICE_ROLE_KEY ใน environment")
    sys.exit(1)

def buddhist_to_ad(date_str):
    """27/09/2565 → 2022-09-27"""
    if not date_str: return None
    if isinstance(date_str, datetime): return date_str.strftime("%Y-%m-%d")
    try:
        d, m, y = str(date_str).split("/")
        return f"{int(y)-543}-{m.zfill(2)}-{d.zfill(2)}"
    except:
        return None

def map_status(thai_status):
    mapping = {
        "ว่าง": "available",
        "ใช้งาน": "in_use",
        "รอดำเนินการ": "pending",
        "ส่งซ่อม": "under_repair",
        "รอทำลาย": "retired",
    }
    return mapping.get(thai_status.strip() if thai_status else "", "available")

def import_data(excel_path: str):
    print(f"📂 กำลังอ่าน: {excel_path}")
    wb = openpyxl.load_workbook(excel_path)
    ws = wb['Sheet 1']

    sp: Client = create_client(SUPABASE_URL, SUPABASE_KEY)

    # --- Step 1: upsert vendors ---
    vendor_names = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[7]:
            vendor_names.add(row[7].strip())

    vendor_map = {}
    for name in vendor_names:
        code = "".join([c for c in name if c.isupper()])[:10] or "VENDOR"
        existing = sp.table("vendors").select("id").eq("name", name).execute()
        if existing.data:
            vendor_map[name] = existing.data[0]["id"]
            print(f"  🏢 {name} → มีอยู่แล้ว")
        else:
            res = sp.table("vendors").insert({"name": name, "code": code}).execute()
            vendor_map[name] = res.data[0]["id"]
            print(f"  🏢 {name} → สร้างใหม่ (code: {code})")

    # --- Step 2: upsert assets ---
    count = 0
    skipped = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        serial, _, com_name, mac_lan, mac_wlan, model, rent, vendor_name, location, status, desc, start, end = row

        if not com_name:
            skipped += 1
            continue

        com_name = com_name.strip()
        asset = {
            "asset_code": com_name,
            "serial_number": serial.strip() if serial else None,
            "model": model.strip() if model else None,
            "mac_lan": mac_lan.strip() if mac_lan else None,
            "mac_wlan": mac_wlan.strip() if mac_wlan else None,
            "vendor_id": vendor_map.get(vendor_name.strip()) if vendor_name else None,
            "monthly_rent": float(rent) if rent else None,
            "location": location.strip() if location else None,
            "status": map_status(status),
            "description": str(desc).strip()[:500] if desc else None,
            "contract_start": buddhist_to_ad(start),
            "contract_end": buddhist_to_ad(end),
        }

        existing = sp.table("assets").select("id").eq("asset_code", com_name).execute()
        if existing.data:
            sp.table("assets").update(asset).eq("asset_code", com_name).execute()
        else:
            sp.table("assets").insert(asset).execute()
        count += 1

    print(f"\n✅ เสร็จสิ้น: {count} เครื่อง, {skipped} ข้าม (ไม่มี com name)")
    print(f"   Vendors: {len(vendor_map)} บริษัท")

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("วิธีใช้: python3 scripts/import_assets.py <path/to/excel.xlsx>")
        sys.exit(1)
    import_data(sys.argv[1])
